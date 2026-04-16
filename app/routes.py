from flask import render_template, redirect, url_for, request, flash, current_app, send_from_directory
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import os
from . import db, csrf
from .models import TVModel, TVModelPhoto, TVModelFirmware, Brand, LauncherType, User, RemoteControl, Tag, AuditLog, ModelComment
from datetime import datetime

ALLOWED_PHOTO    = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_FIRMWARE = {'bin', 'zip', 'img', 'hex'}

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def save_file(file, folder, allowed_extensions):
    if file and file.filename and allowed_file(file.filename, allowed_extensions):
        original_name = secure_filename(file.filename)
        name, ext = os.path.splitext(original_name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
        filename = f"{name}_{timestamp}{ext}"
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], folder, filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        file.save(filepath)
        return filename
    return None

def delete_file(filename, folder):
    if filename:
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], folder, filename)
        if os.path.exists(filepath):
            os.remove(filepath)

def log_action(action, tv_model, field=None, old_value=None, new_value=None):
    """Записывает действие в журнал изменений"""
    try:
        entry = AuditLog(
            tv_model_id=tv_model.id if tv_model else None,
            user_id=current_user.id,
            action=action,
            model_name=tv_model.model if tv_model else None,
            field=field,
            old_value=str(old_value) if old_value is not None else None,
            new_value=str(new_value) if new_value is not None else None,
        )
        db.session.add(entry)
    except Exception:
        pass  # лог не должен ломать основную логику

def editor_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_editor:
            flash('Недостаточно прав для этого действия', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def init_app(app):

    from datetime import timedelta
    import os

    # Состояние импорта — dict чтобы работало в closure и threads
    import os as _os
    _IMPORT_TS_FILE = _os.path.join(_os.path.expanduser('~'), '.kabinet_technologa', 'last_import')

    def _read_last_import():
        try:
            with open(_IMPORT_TS_FILE) as f:
                from datetime import datetime, timezone
                return datetime.fromtimestamp(float(f.read().strip()), tz=timezone.utc)
        except Exception:
            return None

    def _write_last_import(dt):
        try:
            _os.makedirs(_os.path.dirname(_IMPORT_TS_FILE), exist_ok=True)
            with open(_IMPORT_TS_FILE, 'w') as f:
                f.write(str(dt.timestamp()))
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f'_write_last_import error: {e}')

    _state = {'import_running': False, 'last_import_time': _read_last_import()}

    @app.template_filter('local_dt')
    def local_dt(dt):
        if not dt:
            return '—'
        return (dt + timedelta(hours=3)).strftime('%d.%m.%Y %H:%M')

    @app.route('/')
    @login_required
    def index():
        from sqlalchemy import func
        from sqlalchemy.orm import joinedload
        counts = dict(
            db.session.query(TVModel.brand_id, func.count(TVModel.id))
            .group_by(TVModel.brand_id).all()
        )
        brands = Brand.query.order_by(Brand.name).all()
        recent = TVModel.query.options(
            joinedload(TVModel.brand),
        ).filter(TVModel.software_version.isnot(None)).order_by(TVModel.date_added.desc()).limit(10).all()
        import_minutes_ago = None
        if _state['last_import_time']:
            from datetime import datetime, timezone
            import_minutes_ago = int((datetime.now(timezone.utc) - _state['last_import_time']).total_seconds() / 60)
        return render_template('index.html', brands=brands, counts=counts, recent=recent, import_minutes_ago=import_minutes_ago)

    @app.route('/recent-widget')
    @login_required
    def recent_widget():
        from sqlalchemy.orm import joinedload
        models = TVModel.query.options(
            joinedload(TVModel.brand),
            joinedload(TVModel.launcher_type),
        ).order_by(TVModel.date_added.desc()).limit(10).all()
        return render_template('recent_widget.html', models=models)

    @app.route('/brand/<int:brand_id>')
    @login_required
    def brand_detail(brand_id):
        brand = Brand.query.get_or_404(brand_id)
        launcher_types = db.session.query(LauncherType).join(TVModel)\
            .filter(TVModel.brand_id == brand_id).distinct().all()
        return render_template('brand.html', brand=brand, launcher_types=launcher_types)

    @app.route('/brand/<int:brand_id>/launcher/<int:launcher_id>')
    @login_required
    def models_list(brand_id, launcher_id):
        brand    = Brand.query.get_or_404(brand_id)
        launcher = LauncherType.query.get_or_404(launcher_id)

        sort  = request.args.get('sort', 'date')
        order = request.args.get('order', 'desc')
        q     = request.args.get('q', '').strip()
        page  = request.args.get('page', 1, type=int)

        from sqlalchemy import func, cast, Integer

        # Для лота — числовая сортировка (лот может быть "1", "2", "11", "2/2")
        if sort == 'lot':
            try:
                db_url = str(db.engine.url)
                if 'postgresql' in db_url or 'postgres' in db_url:
                    lot_num = func.cast(
                        func.regexp_replace(TVModel.lot, r'[^0-9].*', '', 'g'),
                        Integer
                    )
                else:
                    # SQLite — сортируем как строку
                    lot_num = TVModel.lot
                sort_expr = lot_num.asc() if order == 'asc' else lot_num.desc()
            except Exception:
                sort_expr = TVModel.lot.asc() if order == 'asc' else TVModel.lot.desc()
        else:
            sort_map = {
                'model':  TVModel.model,
                'tester': TVModel.tester_name,
                'date':   TVModel.date_added,
            }
            sort_col  = sort_map.get(sort, TVModel.date_added)
            sort_expr = sort_col.asc() if order == 'asc' else sort_col.desc()

        from sqlalchemy.orm import joinedload
        query = TVModel.query.options(
            joinedload(TVModel.tags),
            joinedload(TVModel.remote),
            joinedload(TVModel.tester),
        ).filter_by(brand_id=brand_id, launcher_type_id=launcher_id)
        if q:
            query = query.filter(TVModel.model.ilike(f'%{q}%'))

        # Фильтр по тегу
        tag_filter = request.args.get('tag', type=int)
        if tag_filter:
            query = query.filter(TVModel.tags.any(Tag.id == tag_filter))

        all_tags = Tag.query.order_by(Tag.name).all()
        all_remotes = RemoteControl.query.order_by(RemoteControl.name).all()
        pagination = query.order_by(sort_expr).paginate(page=page, per_page=20, error_out=False)
        return render_template('models.html',
            brand=brand, launcher=launcher,
            models=pagination.items, pagination=pagination,
            sort=sort, order=order, q=q,
            all_tags=all_tags, tag_filter=tag_filter,
            all_remotes=all_remotes)

    @app.route('/add', methods=['GET', 'POST'])
    @login_required
    @editor_required
    def add_model():
        if request.method == 'POST':
            brand_id       = request.form.get('brand_id')
            launcher_id    = request.form.get('launcher_type_id')
            model          = request.form.get('model')
            lot            = request.form.get('lot')
            specs          = request.form.get('specifications')
            remote         = request.form.get('remote_control')  # это ID пульта
            sw_version     = request.form.get('software_version')   # ← НОВОЕ
            tester_name    = request.form.get('tester_name') or current_user.name
            flashable      = request.form.get('is_flashable') == 'on'

            if not brand_id or not launcher_id or not model or not lot:
                flash('Заполните все обязательные поля', 'error')
                return redirect(url_for('add_model'))

            existing = TVModel.query.filter_by(brand_id=brand_id, model=model, lot=lot).first()
            if existing:
                flash('Модель с таким названием и лотом уже существует', 'error')
                return redirect(url_for('add_model'))

            firmware = request.files.get('firmware')
            firmware_filename = save_file(firmware, 'firmware', ALLOWED_FIRMWARE)

            tv = TVModel(
                brand_id=brand_id, launcher_type_id=launcher_id,
                model=model, lot=lot, specifications=specs,
                remote_control_id=int(remote) if remote else None,
                software_version=sw_version,
                tester_name=tester_name,
                tester_id=current_user.id,
                is_flashable=flashable,
            )
            db.session.add(tv)
            db.session.flush()

            photos = request.files.getlist('photos')
            for i, photo in enumerate(photos):
                fname = save_file(photo, 'photos', ALLOWED_PHOTO)
                if fname:
                    db.session.add(TVModelPhoto(tv_model_id=tv.id, filename=fname, order=i))

            # Несколько прошивок
            for fw_file in request.files.getlist('firmwares'):
                fname = save_file(fw_file, 'firmware', ALLOWED_FIRMWARE)
                if fname:
                    db.session.add(TVModelFirmware(
                        tv_model_id=tv.id,
                        filename=fname,
                        original_name=secure_filename(fw_file.filename)
                    ))

            # Теги
            tag_ids = request.form.getlist('tags')
            tv.tags = Tag.query.filter(Tag.id.in_([int(i) for i in tag_ids if i])).all()

            log_action('create', tv)
            db.session.commit()
            flash('Модель успешно добавлена', 'success')
            back_url = request.form.get('back_url', '')
            if back_url:
                return redirect(back_url)
            return redirect(url_for('models_list', brand_id=brand_id, launcher_id=launcher_id))

        brands = Brand.query.order_by(Brand.name).all()
        launcher_types = LauncherType.query.order_by(LauncherType.name).all()
        remotes = RemoteControl.query.order_by(RemoteControl.name).all()
        all_tags = Tag.query.order_by(Tag.name).all()
        prefill = {
            'brand_id':    request.args.get('prefill_brand', ''),
            'launcher_id': request.args.get('prefill_launcher', ''),
            'model':       request.args.get('prefill_model', ''),
            'lot':         request.args.get('prefill_lot', ''),
            'remote':      request.args.get('prefill_remote', ''),
            'sw_version':  request.args.get('prefill_sw', ''),
            'tester':      request.args.get('prefill_tester', ''),
            'flashable':   request.args.get('prefill_flashable', '0') == '1',
            'specs':       request.args.get('prefill_specs', ''),
        }
        back_url = request.args.get('back_url', '')

        # Подгружаем объекты для хлебных крошек
        prefill_brand_obj    = Brand.query.get(prefill['brand_id'])    if prefill['brand_id']    else None
        prefill_launcher_obj = LauncherType.query.get(prefill['launcher_id']) if prefill['launcher_id'] else None

        return render_template('add.html', brands=brands, launcher_types=launcher_types,
                               remotes=remotes, all_tags=all_tags, prefill=prefill, back_url=back_url,
                               prefill_brand_obj=prefill_brand_obj, prefill_launcher_obj=prefill_launcher_obj)

    @app.route('/view/<int:id>')
    @login_required
    def view_model(id):
        from sqlalchemy.orm import joinedload
        model = TVModel.query.options(
            joinedload(TVModel.brand),
            joinedload(TVModel.launcher_type),
            joinedload(TVModel.photos),
            joinedload(TVModel.firmwares),
            joinedload(TVModel.tags),
            joinedload(TVModel.remote),
        ).get_or_404(id)
        back_url = request.args.get('back_url') or request.referrer or \
                   url_for('models_list', brand_id=model.brand_id, launcher_id=model.launcher_type_id)
        audit_log = AuditLog.query.filter_by(tv_model_id=id)\
            .order_by(AuditLog.timestamp.desc()).limit(30).all()
        comments = ModelComment.query.filter_by(tv_model_id=id)\
            .order_by(ModelComment.timestamp.asc()).all()
        return render_template('view.html', model=model, back_url=back_url,
                               audit_log=audit_log, comments=comments)

    @app.route('/edit/<int:id>', methods=['GET', 'POST'])
    @login_required
    @editor_required
    def edit_model(id):
        model = TVModel.query.get_or_404(id)

        if request.method == 'POST':
            # Снимаем старые значения ДО изменения
            _old = {
                'model':            model.model,
                'lot':              model.lot,
                'software_version': model.software_version,
                'tester_name':      model.tester_name,
                'is_flashable':     model.is_flashable,
                'remote_control_id':model.remote_control_id,
                'specifications':   model.specifications,
            }

            model.brand_id         = request.form.get('brand_id')
            model.launcher_type_id = request.form.get('launcher_type_id')
            model.model            = request.form.get('model')
            model.lot              = request.form.get('lot')
            model.specifications   = request.form.get('specifications')
            model.remote_control_id = int(request.form.get('remote_control')) if request.form.get('remote_control') else None
            model.software_version = request.form.get('software_version')
            model.tester_name      = request.form.get('tester_name') or current_user.name
            model.is_flashable     = request.form.get('is_flashable') == 'on'

            if not all([model.brand_id, model.launcher_type_id, model.model, model.lot]):
                flash('Заполните все обязательные поля', 'error')
                return redirect(url_for('edit_model', id=id))

            duplicate = TVModel.query.filter(
                TVModel.brand_id == model.brand_id,
                TVModel.model == model.model,
                TVModel.lot == model.lot,
                TVModel.id != id
            ).first()
            if duplicate:
                flash('Модель с таким названием и лотом уже существует', 'error')
                return redirect(url_for('edit_model', id=id))

            # Удаление отмеченных фото
            for pid in request.form.getlist('delete_photos'):
                photo = TVModelPhoto.query.get(int(pid))
                if photo and photo.tv_model_id == id:
                    delete_file(photo.filename, 'photos')
                    db.session.delete(photo)

            # Новые фото
            current_count = TVModelPhoto.query.filter_by(tv_model_id=id).count()
            for i, photo in enumerate(request.files.getlist('photos')):
                fname = save_file(photo, 'photos', ALLOWED_PHOTO)
                if fname:
                    db.session.add(TVModelPhoto(tv_model_id=id, filename=fname, order=current_count + i))

            # Удаление отмеченных прошивок
            for fid in request.form.getlist('delete_firmwares'):
                fw = TVModelFirmware.query.get(int(fid))
                if fw and fw.tv_model_id == id:
                    delete_file(fw.filename, 'firmware')
                    db.session.delete(fw)

            # Новые прошивки
            for fw_file in request.files.getlist('firmwares'):
                fname = save_file(fw_file, 'firmware', ALLOWED_FIRMWARE)
                if fname:
                    db.session.add(TVModelFirmware(
                        tv_model_id=id,
                        filename=fname,
                        original_name=secure_filename(fw_file.filename)
                    ))

            # Теги
            tag_ids = request.form.getlist('tags')
            model.tags = Tag.query.filter(Tag.id.in_([int(i) for i in tag_ids if i])).all()

            # Лог изменений — сравниваем старые и новые значения
            _new = {
                'model':            model.model,
                'lot':              model.lot,
                'software_version': model.software_version,
                'tester_name':      model.tester_name,
                'is_flashable':     model.is_flashable,
                'remote_control_id':model.remote_control_id,
                'specifications':   model.specifications,
            }
            for field_key, old_v in _old.items():
                new_v = _new[field_key]
                if str(old_v or '') != str(new_v or ''):
                    log_action('edit', model, field=field_key, old_value=old_v, new_value=new_v)

            db.session.commit()
            flash('Модель обновлена', 'success')
            back_url = request.form.get('back_url', '')
            return redirect(url_for('edit_model', id=id, back_url=back_url))

        back_url = request.args.get('back_url') or request.referrer or \
                   url_for('models_list', brand_id=model.brand_id, launcher_id=model.launcher_type_id)

        brands = Brand.query.order_by(Brand.name).all()
        launcher_types = LauncherType.query.order_by(LauncherType.name).all()
        all_tags = Tag.query.order_by(Tag.name).all()
        remotes = RemoteControl.query.order_by(RemoteControl.name).all()
        return render_template('edit.html', model=model, brands=brands,
                               launcher_types=launcher_types, remotes=remotes,
                               all_tags=all_tags, back_url=back_url)

    @app.route('/duplicate/<int:id>')
    @login_required
    @editor_required
    def duplicate_model(id):
        src = TVModel.query.get_or_404(id)
        # Не создаём запись — просто открываем форму добавления с предзаполненными данными
        return redirect(url_for('add_model',
            prefill_brand=src.brand_id,
            prefill_launcher=src.launcher_type_id,
            prefill_model=src.model + ' (копия)',
            prefill_lot=src.lot,
            prefill_remote=src.remote.name if src.remote else '',
            prefill_sw=src.software_version or '',
            prefill_tester=src.tester_name or '',
            prefill_flashable='1' if src.is_flashable else '0',
            prefill_specs=src.specifications or '',
            back_url=url_for('models_list', brand_id=src.brand_id, launcher_id=src.launcher_type_id),
        ))

    @app.route('/delete/<int:id>', methods=['POST'])
    @login_required
    @editor_required
    def delete_model(id):
        model = TVModel.query.get_or_404(id)
        brand_id, launcher_id = model.brand_id, model.launcher_type_id
        log_action('delete', model)
        for photo in model.photos:
            delete_file(photo.filename, 'photos')
        for fw in model.firmwares:
            delete_file(fw.filename, 'firmware')
        if model.firmware_filename:
            delete_file(model.firmware_filename, 'firmware')
        db.session.delete(model)
        db.session.commit()
        flash('Модель удалена', 'success')
        return redirect(url_for('models_list', brand_id=brand_id, launcher_id=launcher_id))

    @app.route('/download/<int:id>/<file_type>')
    @login_required
    def download_file(id, file_type):
        model = TVModel.query.get_or_404(id)
        if file_type == 'firmware' and model.firmware_filename:
            return send_from_directory(
                os.path.join(current_app.config['UPLOAD_FOLDER'], 'firmware'),
                model.firmware_filename, as_attachment=True,
                download_name=model.firmware_filename)
        flash('Файл не найден', 'error')
        return redirect(url_for('view_model', id=id))

    @app.route('/download_firmware/<int:fw_id>')
    @login_required
    def download_firmware(fw_id):
        fw = TVModelFirmware.query.get_or_404(fw_id)
        return send_from_directory(
            os.path.join(current_app.config['UPLOAD_FOLDER'], 'firmware'),
            fw.filename, as_attachment=True,
            download_name=fw.original_name)

    @app.route('/delete_firmware/<int:fw_id>', methods=['POST'])
    @login_required
    @editor_required
    def delete_firmware(fw_id):
        fw = TVModelFirmware.query.get_or_404(fw_id)
        model_id = fw.tv_model_id
        delete_file(fw.filename, 'firmware')
        db.session.delete(fw)
        db.session.commit()
        flash('Прошивка удалена', 'success')
        return redirect(url_for('view_model', id=model_id))

    @app.route('/desktop-autologin')
    def desktop_autologin():
        from flask_login import login_user, current_user
        from flask import redirect, url_for
        import logging
        logger = logging.getLogger(__name__)
        if current_user.is_authenticated:
            return redirect(url_for('index'))
        try:
            import os, json, hashlib, base64, platform
            from cryptography.fernet import Fernet
            home = os.environ.get('APPDATA') or os.environ.get('USERPROFILE') or os.path.expanduser('~')
            creds_file = os.path.join(home, 'KabinetTechnologa', 'creds')
            logger.warning(f'AUTOLOGIN: creds_file={creds_file}, exists={os.path.exists(creds_file)}')
            if os.path.exists(creds_file):
                fingerprint = f"{platform.node()}:{os.environ.get('USERNAME') or os.environ.get('USER') or 'user'}"
                key = base64.urlsafe_b64encode(hashlib.sha256(fingerprint.encode()).digest())
                f = Fernet(key)
                with open(creds_file, 'rb') as fp:
                    data = json.loads(f.decrypt(fp.read()).decode())
                user = User.query.filter_by(email=data['email'].lower()).first()
                logger.warning(f'AUTOLOGIN: user={user}, email={data["email"]}')
                if user and user.check_password(data['password']) and user.is_active_user:
                    login_user(user, remember=True)
                    logger.warning('AUTOLOGIN: success')
                    return redirect(url_for('index'))
        except Exception as e:
            logger.warning(f'AUTOLOGIN error: {e}')
        return redirect(url_for('auth.login'))

    @app.route('/api/import-status')
    @login_required
    def import_status():
        from flask import jsonify
        from datetime import datetime, timezone
        if _state['last_import_time']:
            minutes = int((datetime.now(timezone.utc) - _state['last_import_time']).total_seconds() / 60)
            return jsonify({'ok': True, 'minutes_ago': minutes})
        return jsonify({'ok': False})

    @app.route('/api/suggest')
    @login_required
    def suggest():
        from flask import jsonify
        q = request.args.get('q', '').strip()
        brand_id = request.args.get('brand_id', '').strip()
        if len(q) < 2:
            return jsonify([])

        words = q.split()
        from sqlalchemy.orm import joinedload
        query = TVModel.query.join(Brand).join(LauncherType).options(
            joinedload(TVModel.brand),
            joinedload(TVModel.launcher_type),
        )

        # Фильтруем по бренду если передан (локальный поиск)
        if brand_id:
            query = query.filter(TVModel.brand_id == brand_id)

        # Каждое слово должно встречаться хоть в одном поле
        for word in words:
            query = query.filter(db.or_(
                TVModel.model.ilike(f'%{word}%'),
                TVModel.lot.ilike(f'%{word}%'),
                TVModel.software_version.ilike(f'%{word}%'),
                Brand.name.ilike(f'%{word}%'),
            ))

        results = query.order_by(TVModel.date_added.desc()).limit(8).all()
        return jsonify([{
            'id':       m.id,
            'model':    m.model,
            'brand':    m.brand.name,
            'launcher': m.launcher_type.name,
            'lot':      m.lot,
            'sw':       m.software_version or '',
            'url':      url_for('view_model', id=m.id),
        } for m in results])

    # ── СИНХРОНИЗАЦИЯ С GOOGLE SHEETS ──
    @app.route('/sync/sheets', methods=['POST'])
    @login_required
    @editor_required
    def sync_sheets():
        import os
        try:
            import gspread
            from google.oauth2.service_account import Credentials
        except ImportError:
            flash('Установите: pip install gspread google-auth', 'error')
            return redirect(url_for('import_all'))

        creds_file = current_app.config.get('SHEETS_CREDENTIALS_FILE')
        sheet_id   = current_app.config.get('SHEETS_SPREADSHEET_ID')

        if not creds_file or not os.path.exists(creds_file):
            flash('Файл google_credentials.json не найден в корне проекта', 'error')
            return redirect(url_for('import_all'))

        try:
            scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly']
            creds  = Credentials.from_service_account_file(creds_file, scopes=scopes)
            gc     = gspread.authorize(creds)
            wb     = gc.open_by_key(sheet_id)
        except Exception as e:
            flash(f'Ошибка подключения к Google Sheets: {e}', 'error')
            return redirect(url_for('import_all'))

        launcher_name = 'собственный'
        launcher = LauncherType.query.filter_by(name=launcher_name).first()
        if not launcher:
            launcher = LauncherType(name=launcher_name)
            db.session.add(launcher)
            db.session.flush()

        brand_cache  = {b.name: b for b in Brand.query.all()}
        remote_cache = {r.name: r for r in RemoteControl.query.all()}
        existing_map = {
            (m.brand_id, m.model, m.lot): m.id
            for m in db.session.query(TVModel.brand_id, TVModel.model, TVModel.lot, TVModel.id).all()
        }
        existing_set = set(existing_map.keys())

        skip_sheets = {'Требования по качеству'}
        added = skipped = 0

        try:
            for sheet in wb.worksheets():
                if sheet.title in skip_sheets:
                    continue

                brand_name = sheet.title.strip()
                if brand_name not in brand_cache:
                    brand = Brand(name=brand_name)
                    db.session.add(brand)
                    db.session.flush()
                    brand_cache[brand_name] = brand
                brand = brand_cache[brand_name]

                all_rows = sheet.get_all_values()
                if len(all_rows) < 4:
                    continue

                # Определяем колонки по заголовкам строк 3-4
                col_tester = 7; col_flash = 8; col_sw = 9; col_remote = 10
                for hi in [2, 3]:
                    if hi >= len(all_rows):
                        continue
                    for ci, val in enumerate(all_rows[hi]):
                        v = val.strip().lower()
                        if 'разработчик ртп' in v: col_tester = ci
                        elif v in ('шьём', 'шьем'): col_flash = ci
                        elif 'версия по' in v: col_sw = ci
                        elif v == 'stb': col_remote = ci

                def _gs(r, idx):
                    return r[idx].strip() if len(r) > idx else ''

                for row in all_rows[4:]:
                    if not row:
                        continue
                    model_name = _gs(row, 0)
                    if not model_name or model_name == 'None':
                        continue

                    lot_raw = _gs(row, 1)
                    if not lot_raw:
                        continue
                    try:
                        lot = str(int(float(lot_raw))) if '.' in lot_raw and lot_raw.replace('.', '').isdigit() else lot_raw
                    except Exception:
                        lot = lot_raw

                    tester    = _gs(row, col_tester)
                    flashable = _gs(row, col_flash).lower() == 'да'
                    sw        = _gs(row, col_sw)
                    remote    = _gs(row, col_remote)

                    remote_id = None
                    if remote:
                        if remote not in remote_cache:
                            rc = RemoteControl(name=remote)
                            db.session.add(rc)
                            db.session.flush()
                            remote_cache[remote] = rc
                        remote_id = remote_cache[remote].id

                    if (brand.id, model_name, lot) in existing_set:
                        tv_id = existing_map.get((brand.id, model_name, lot))
                        if tv_id:
                            tv = TVModel.query.get(tv_id)
                            if tv:
                                if sw: tv.software_version = sw
                                if tester: tv.tester_name = tester
                                if remote_id: tv.remote_control_id = remote_id
                                tv.is_flashable = flashable
                        skipped += 1
                        continue

                    tv = TVModel(
                        brand_id=brand.id,
                        launcher_type_id=launcher.id,
                        model=model_name, lot=lot,
                        remote_control_id=remote_id,
                        software_version=sw or None,
                        is_flashable=flashable,
                        tester_name=tester or None,
                        tester_id=current_user.id if tester else None,
                    )
                    db.session.add(tv)
                    existing_set.add((brand.id, model_name, lot))
                    log_action('create', tv)
                    added += 1

                    if added % 100 == 0:
                        db.session.commit()

            db.session.commit()
            flash(f'Синхронизация завершена: добавлено {added}, пропущено дублей {skipped}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка синхронизации: {e}', 'error')

        return redirect(url_for('index'))

    # ── КОММЕНТАРИИ К МОДЕЛИ ──
    @app.route('/comment/<int:model_id>', methods=['POST'])
    @login_required
    @editor_required
    @csrf.exempt
    def add_comment(model_id):
        from flask import jsonify
        model = TVModel.query.get_or_404(model_id)
        text = request.get_json().get('text', '').strip()
        if not text:
            return jsonify({'ok': False, 'error': 'Пустой комментарий'}), 400
        comment = ModelComment(
            tv_model_id=model_id,
            user_id=current_user.id,
            text=text
        )
        db.session.add(comment)
        db.session.commit()
        return jsonify({
            'ok': True,
            'id': comment.id,
            'text': comment.text,
            'user': current_user.name,
            'time': comment.timestamp.strftime('%d.%m.%Y %H:%M'),
            'can_delete': True
        })

    @app.route('/comment/<int:comment_id>/delete', methods=['POST'])
    @login_required
    @csrf.exempt
    def delete_comment(comment_id):
        from flask import jsonify
        comment = ModelComment.query.get_or_404(comment_id)
        # Удалить может автор или админ
        if comment.user_id != current_user.id and not current_user.is_admin:
            return jsonify({'ok': False, 'error': 'Нет прав'}), 403
        db.session.delete(comment)
        db.session.commit()
        return jsonify({'ok': True})

    @app.route('/search')
    @login_required
    def search():
        q = request.args.get('q', '').strip()
        results = []
        if q:
            words = q.split()
            query = TVModel.query.join(Brand).join(LauncherType)
            for word in words:
                query = query.filter(db.or_(
                    TVModel.model.ilike(f'%{word}%'),
                    TVModel.lot.ilike(f'%{word}%'),
                    TVModel.tester_name.ilike(f'%{word}%'),
                    TVModel.software_version.ilike(f'%{word}%'),
                    Brand.name.ilike(f'%{word}%'),
                ))
            results = query.order_by(TVModel.date_added.desc()).limit(50).all()
        return render_template('search.html', q=q, results=results)

    @app.route('/download_photo/<int:photo_id>')
    @login_required
    def download_photo(photo_id):
        photo = TVModelPhoto.query.get_or_404(photo_id)
        return send_from_directory(
            os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos'),
            photo.filename, as_attachment=True, download_name=photo.filename)

    # ── ГЛОБАЛЬНЫЙ ЭКСПОРТ ВСЕГО ──
    @app.route('/export/all')
    @login_required
    def export_all():
        import io, csv
        from flask import Response
        from sqlalchemy.orm import joinedload
        models = TVModel.query.join(Brand).join(LauncherType).options(
            joinedload(TVModel.brand),
            joinedload(TVModel.launcher_type),
            joinedload(TVModel.remote),
            joinedload(TVModel.tags),
        ).order_by(Brand.name, LauncherType.name, TVModel.model).all()

        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow([
            'Бренд', 'Лаунчер', 'Модель', 'Лот', 'Пульт', 'Версия ПО',
            'Прошивается', 'Тестировщик', 'Метки', 'Дата добавления'
        ])
        for m in models:
            writer.writerow([
                m.brand.name,
                m.launcher_type.name,
                m.model,
                m.lot,
                m.remote.name if m.remote else '',
                m.software_version or '',
                'Да' if m.is_flashable else 'Нет',
                m.tester_name or '',
                ', '.join(t.name for t in m.tags),
                m.date_added.strftime('%d.%m.%Y'),
            ])

        output.seek(0)
        return Response(
            '\ufeff' + output.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename="kabinet_export.csv"'}
        )


    # ── АВТОИМПОРТ ИЗ APPS SCRIPT ──
    @app.route('/api/auto-import', methods=['POST'])
    @csrf.exempt
    def auto_import():
        from flask import jsonify
        import io, threading, openpyxl, logging
        logger = logging.getLogger(__name__)

        secret = current_app.config.get('IMPORT_SECRET', '')
        if not secret:
            return jsonify({'ok': False, 'error': 'IMPORT_SECRET не настроен'}), 500

        token = request.headers.get('X-Import-Token') or request.form.get('token', '')
        if token != secret:
            return jsonify({'ok': False, 'error': 'Неверный токен'}), 403

        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'ok': False, 'error': 'Файл не передан'}), 400

        if file.filename.rsplit('.', 1)[-1].lower() not in ('xlsx', 'xls'):
            return jsonify({'ok': False, 'error': 'Только xlsx/xls'}), 400

        file_bytes = io.BytesIO(file.read())
        app_obj = current_app._get_current_object()

        def do_import():
            with app_obj.app_context():
                try:
                    db.session.rollback()
                    logger.warning('AUTO_IMPORT: opening workbook...')
                    wb = openpyxl.load_workbook(file_bytes, data_only=True, read_only=True)
                    logger.warning(f'AUTO_IMPORT: workbook opened, sheets: {wb.sheetnames[:5]}')

                    # Читаем threaded comments напрямую из XML внутри xlsx
                    logger.warning('AUTO_IMPORT: extracting threaded comments...')
                    comments_from_cells = {}
                    dates_map = {}
                    try:
                        import zipfile as zf_mod
                        from xml.etree import ElementTree as ET
                        TC_NS = 'http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments'
                        WB_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

                        file_bytes.seek(0)
                        zf = zf_mod.ZipFile(file_bytes)
                        wb_xml = ET.fromstring(zf.read('xl/workbook.xml'))
                        sheets_order = [s.get('name') for s in wb_xml.findall(f'.//{{{WB_NS}}}sheet')]

                        for i, sname in enumerate(sheets_order, start=1):
                            rels_path = f'xl/worksheets/_rels/sheet{i}.xml.rels'
                            if rels_path not in zf.namelist():
                                continue
                            rels_xml = ET.fromstring(zf.read(rels_path))
                            tc_file = None
                            for r in rels_xml:
                                if 'threadedComment' in r.get('Type', ''):
                                    tc_file = 'xl/threadedComments/' + r.get('Target', '').split('/')[-1]
                                    break
                            if not tc_file or tc_file not in zf.namelist():
                                continue
                            tc_xml = ET.fromstring(zf.read(tc_file))
                            seen_refs = set()
                            for tc in tc_xml.findall(f'{{{TC_NS}}}threadedComment'):
                                if tc.get('parentId'):
                                    continue
                                ref = tc.get('ref', '')
                                col = ''.join(c for c in ref if c.isalpha())
                                if col != 'J':
                                    continue
                                if ref in seen_refs:
                                    continue
                                row = int(''.join(c for c in ref if c.isdigit()))
                                text_el = tc.find(f'{{{TC_NS}}}text')
                                if text_el is not None and text_el.text:
                                    comments_from_cells[(sname, row)] = text_el.text.strip()
                                    seen_refs.add(ref)
                                # Дата из dT атрибута
                                dt_str = tc.get('dT', '')
                                if dt_str:
                                    try:
                                        from datetime import datetime
                                        dt = datetime.fromisoformat(dt_str.rstrip('0').rstrip('.').replace('Z',''))
                                        dates_map[(sname, row)] = dt
                                    except Exception:
                                        pass
                        zf.close()
                        logger.warning(f'AUTO_IMPORT: found {len(comments_from_cells)} threaded comments')
                    except Exception as ce:
                        logger.warning(f'AUTO_IMPORT: comment extraction failed: {ce}')

                    launcher_name = 'собственный'
                    launcher = LauncherType.query.filter_by(name=launcher_name).first()
                    if not launcher:
                        launcher = LauncherType(name=launcher_name)
                        db.session.add(launcher)
                        db.session.flush()

                    logger.warning('AUTO_IMPORT: loading caches...')
                    brand_cache  = {b.name: b for b in Brand.query.all()}
                    remote_cache = {r.name: r for r in RemoteControl.query.all()}
                    existing_map = {
                        (m.brand_id, m.model, m.lot): {
                            'id': m.id,
                            'software_version': m.software_version,
                            'specifications': m.specifications,
                            'tester_name': m.tester_name,
                            'remote_control_id': m.remote_control_id,
                            'is_flashable': m.is_flashable,
                            'date_added': m.date_added,
                        }
                        for m in db.session.query(
                            TVModel.brand_id, TVModel.model, TVModel.lot, TVModel.id,
                            TVModel.software_version, TVModel.specifications,
                            TVModel.tester_name, TVModel.remote_control_id, TVModel.is_flashable,
                            TVModel.date_added
                        ).all()
                    }
                    existing_set = set(existing_map.keys())
                    logger.warning(f'AUTO_IMPORT: caches loaded, existing models: {len(existing_set)}')

                    skip_sheets = {'Требования по качеству'}
                    BATCH = 500
                    added = skipped = 0
                    updates_list = []
                    new_models = []


                    for sheet_name in wb.sheetnames:
                        if sheet_name in skip_sheets:
                            continue
                        logger.warning(f'AUTO_IMPORT: processing sheet {sheet_name}...')
                        brand_name = sheet_name.strip()
                        if brand_name not in brand_cache:
                            brand = Brand(name=brand_name)
                            db.session.add(brand)
                            db.session.flush()
                            brand_cache[brand_name] = brand
                        brand = brand_cache[brand_name]

                        ws = wb[sheet_name]
                        try:
                            tab_color = ws.sheet_properties.tabColor
                            if tab_color and not brand.tab_color:
                                brand.tab_color = tab_color.rgb
                        except Exception:
                            pass

                        col_tester = 7; col_flash = 8; col_sw = 9; col_remote = 10
                        try:
                            for hr in ws.iter_rows(min_row=3, max_row=4, values_only=False):
                                for ci, hc in enumerate(hr):
                                    v = str(hc.value or '').strip().lower()
                                    if 'разработчик ртп' in v: col_tester = ci
                                    elif v in ('шьём', 'шьем'): col_flash = ci
                                    elif 'версия по' in v or 'версия пo' in v: col_sw = ci
                                    elif v == 'stb': col_remote = ci
                        except Exception:
                            pass

                        def _cell_str(r, idx):
                            return str(r[idx].value).strip() if len(r) > idx and r[idx].value else ''

                        for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=False), start=5):
                            model_name = str(row[0].value).strip() if row[0].value else ''
                            if not model_name or model_name == 'None':
                                continue
                            lot_raw = row[1].value
                            if isinstance(lot_raw, float):
                                lot = str(int(lot_raw))
                            elif hasattr(lot_raw, 'strftime'):
                                lot = f"{lot_raw.day}/{lot_raw.month}"
                            else:
                                lot = str(lot_raw or '').strip()
                            if not lot or lot == 'None':
                                continue

                            tester     = _cell_str(row, col_tester)
                            flashable  = _cell_str(row, col_flash).lower() == 'да'
                            sw_version = _cell_str(row, col_sw)
                            remote     = _cell_str(row, col_remote)
                            sw_comment = comments_from_cells.get((sheet_name, row_num), '')

                            remote_id = None
                            if remote:
                                if remote not in remote_cache:
                                    rc = RemoteControl(name=remote)
                                    db.session.add(rc)
                                    db.session.flush()
                                    remote_cache[remote] = rc
                                remote_id = remote_cache[remote].id

                            if (brand.id, model_name, lot) in existing_set:
                                cur = existing_map.get((brand.id, model_name, lot))
                                if cur:
                                    tv_id = cur['id']
                                    # Сравниваем — обновляем только если что-то изменилось
                                    upd = {'id': tv_id}
                                    if sw_version and sw_version != cur['software_version']:
                                        upd['software_version'] = sw_version
                                    if sw_comment and sw_comment != cur['specifications']:
                                        upd['specifications'] = sw_comment
                                    if tester and tester != cur['tester_name']:
                                        upd['tester_name'] = tester
                                    if remote_id and remote_id != cur['remote_control_id']:
                                        upd['remote_control_id'] = remote_id
                                    if bool(flashable) != bool(cur['is_flashable']):
                                        upd['is_flashable'] = flashable
                                    # Обновляем date_added если есть дата из комментария
                                    # и текущая дата выглядит как дата массового импорта
                                    tc_date = dates_map.get((sheet_name, row_num))
                                    if tc_date and cur['date_added']:
                                        # Обновляем если дата из комментария отличается более чем на час
                                        diff = abs((tc_date - cur['date_added']).total_seconds())
                                        if diff > 3600:
                                            upd['date_added'] = tc_date
                                    elif tc_date and not cur['date_added']:
                                        upd['date_added'] = tc_date
                                    if len(upd) > 1:
                                        updates_list.append(upd)
                                skipped += 1
                                continue

                            new_models.append(TVModel(
                                brand_id=brand.id,
                                launcher_type_id=launcher.id,
                                model=model_name, lot=lot,
                                remote_control_id=remote_id,
                                software_version=sw_version or None,
                                specifications=sw_comment or None,
                                is_flashable=flashable,
                                tester_name=tester or None,
                                tester_id=None,
                                date_added=dates_map.get((sheet_name, row_num)),
                            ))
                            existing_set.add((brand.id, model_name, lot))
                            added += 1

                    wb.close()
                    logger.warning(f'AUTO_IMPORT: inserting {added} new, updating {len(updates_list)} changed (skipped {skipped - len(updates_list)} unchanged)...')

                    from sqlalchemy.dialects.postgresql import insert as pg_insert

                    # Upsert новых одним запросом
                    if new_models:
                        rows = [{
                            'brand_id': m.brand_id,
                            'launcher_type_id': m.launcher_type_id,
                            'model': m.model,
                            'lot': m.lot,
                            'remote_control_id': m.remote_control_id,
                            'software_version': m.software_version,
                            'specifications': m.specifications,
                            'is_flashable': m.is_flashable,
                            'tester_name': m.tester_name,
                            'tester_id': m.tester_id,
                        } for m in new_models]
                        stmt = pg_insert(TVModel).values(rows)
                        stmt = stmt.on_conflict_do_nothing()
                        db.session.execute(stmt)
                        db.session.commit()
                        logger.warning(f'AUTO_IMPORT: inserted {len(rows)} new')

                    # Bulk update существующих чанками с retry
                    if updates_list:
                        CHUNK = 500
                        total_chunks = (len(updates_list) - 1) // CHUNK + 1
                        for i in range(0, len(updates_list), CHUNK):
                            chunk = updates_list[i:i+CHUNK]
                            for attempt in range(3):
                                try:
                                    db.session.rollback()
                                    db.session.bulk_update_mappings(TVModel, chunk)
                                    db.session.commit()
                                    logger.warning(f'AUTO_IMPORT: updated chunk {i//CHUNK + 1}/{total_chunks}')
                                    break
                                except Exception as chunk_err:
                                    logger.error(f'AUTO_IMPORT chunk error attempt {attempt+1}: {chunk_err}')
                                    db.session.rollback()
                                    if attempt == 2:
                                        raise

                    logger.warning('AUTO_IMPORT: done!')
                    from datetime import datetime, timezone
                    _state['last_import_time'] = datetime.now(timezone.utc)
                    _write_last_import(_state['last_import_time'])

                except Exception as e:
                    db.session.rollback()
                    logger.error(f'AUTO_IMPORT error: {e}')

        if _state['import_running']:
            return jsonify({'ok': False, 'error': 'Импорт уже запущен, подождите'}), 429

        def do_import_safe():
            _state['import_running'] = True
            try:
                do_import()
            finally:
                _state['import_running'] = False

        t = threading.Thread(target=do_import_safe, daemon=True)
        t.start()
        return jsonify({'ok': True, 'status': 'processing'})
    # ── ГЛОБАЛЬНЫЙ ИМПОРТ ──
    @app.route('/import/all', methods=['GET', 'POST'])
    @login_required
    @editor_required
    def import_all():
        if request.method == 'POST':
            import io
            file = request.files.get('file')
            if not file or not file.filename:
                flash('Выберите файл', 'error')
                return redirect(request.url)

            launcher_name = 'собственный'
            ext = file.filename.rsplit('.', 1)[-1].lower()

            try:
                added = skipped = 0
                BATCH = 100  # коммит каждые 100 записей

                # ── Кэш брендов, пультов, существующих моделей ──
                brand_cache   = {b.name: b for b in Brand.query.all()}
                remote_cache  = {r.name: r for r in RemoteControl.query.all()}
                existing_map  = {
                    (m.brand_id, m.model, m.lot): m.id
                    for m in db.session.query(TVModel.brand_id, TVModel.model, TVModel.lot, TVModel.id).all()
                }
                existing_set  = set(existing_map.keys())

                launcher = LauncherType.query.filter_by(name=launcher_name).first()
                if not launcher:
                    launcher = LauncherType(name=launcher_name)
                    db.session.add(launcher)
                    db.session.flush()

                if ext in ('xlsx', 'xls'):
                    import openpyxl, zipfile as zf_mod
                    from xml.etree import ElementTree as ET
                    file_bytes = io.BytesIO(file.read())

                    # Читаем threaded comments через zipfile
                    tc_comments = {}
                    tc_dates = {}
                    try:
                        TC_NS = 'http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments'
                        WB_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                        zf = zf_mod.ZipFile(file_bytes)
                        wb_xml = ET.fromstring(zf.read('xl/workbook.xml'))
                        sheets_order = [s.get('name') for s in wb_xml.findall(f'.//{{{WB_NS}}}sheet')]
                        for i, sname in enumerate(sheets_order, start=1):
                            rels_path = f'xl/worksheets/_rels/sheet{i}.xml.rels'
                            if rels_path not in zf.namelist():
                                continue
                            rels_xml = ET.fromstring(zf.read(rels_path))
                            tc_file = None
                            for r in rels_xml:
                                if 'threadedComment' in r.get('Type', ''):
                                    tc_file = 'xl/threadedComments/' + r.get('Target', '').split('/')[-1]
                                    break
                            if not tc_file or tc_file not in zf.namelist():
                                continue
                            tc_xml = ET.fromstring(zf.read(tc_file))
                            seen = set()
                            for tc in tc_xml.findall(f'{{{TC_NS}}}threadedComment'):
                                if tc.get('parentId'):
                                    continue
                                ref = tc.get('ref', '')
                                col = ''.join(c for c in ref if c.isalpha())
                                if col != 'J' or ref in seen:
                                    continue
                                row_num = int(''.join(c for c in ref if c.isdigit()))
                                text_el = tc.find(f'{{{TC_NS}}}text')
                                if text_el is not None and text_el.text:
                                    tc_comments[(sname, row_num)] = text_el.text.strip()
                                    seen.add(ref)
                                dt_str = tc.get('dT', '')
                                if dt_str:
                                    try:
                                        from datetime import datetime
                                        dt = datetime.fromisoformat(dt_str.rstrip('0').rstrip('.').replace('Z',''))
                                        tc_dates[(sname, row_num)] = dt
                                    except Exception:
                                        pass
                        zf.close()
                    except Exception:
                        pass

                    file_bytes.seek(0)
                    wb = openpyxl.load_workbook(file_bytes, data_only=True, read_only=True)

                    skip_sheets = {'Требования по качеству'}

                    for sheet_name in wb.sheetnames:
                        if sheet_name in skip_sheets:
                            continue

                        brand_name = sheet_name.strip()
                        if brand_name not in brand_cache:
                            brand = Brand(name=brand_name)
                            db.session.add(brand)
                            db.session.flush()
                            brand_cache[brand_name] = brand
                        brand = brand_cache[brand_name]

                        ws = wb[sheet_name]

                        # Сохраняем цвет вкладки
                        try:
                            tab_color = ws.sheet_properties.tabColor
                            if tab_color and not brand_cache[brand_name].tab_color:
                                brand_cache[brand_name].tab_color = tab_color.rgb
                        except Exception:
                            pass

                        # Определяем индексы колонок по заголовкам строк 3-4
                        col_tester = 7
                        col_flash  = 8
                        col_sw     = 9
                        col_remote = 10
                        try:
                            for hr in ws.iter_rows(min_row=3, max_row=4, values_only=False):
                                for ci, hc in enumerate(hr):
                                    v = str(hc.value or '').strip().lower()
                                    if 'разработчик ртп' in v:
                                        col_tester = ci
                                    elif v == 'шьём' or v == 'шьем':
                                        col_flash = ci
                                    elif 'версия по' in v or 'версия пo' in v:
                                        col_sw = ci
                                    elif v == 'stb':
                                        col_remote = ci
                        except Exception:
                            pass

                        def _cell_str(r, idx):
                            return str(r[idx].value).strip() if len(r) > idx and r[idx].value else ''

                        for row in ws.iter_rows(min_row=5, values_only=False):
                            model_name = str(row[0].value).strip() if row[0].value else ''
                            if not model_name or model_name == 'None':
                                continue
                            lot_raw  = row[1].value
                            lot_cell = row[1]
                            if isinstance(lot_raw, float):
                                lot = str(int(lot_raw))
                            elif hasattr(lot_raw, 'strftime'):
                                fmt = getattr(lot_cell, 'number_format', '') or ''
                                if 'd/m' in fmt.lower():
                                    lot = f"{lot_raw.day}/{lot_raw.month}"
                                elif 'm/d' in fmt.lower():
                                    lot = f"{lot_raw.month}/{lot_raw.day}"
                                else:
                                    lot = f"{lot_raw.day}/{lot_raw.month}"
                            else:
                                lot = str(lot_raw or '').strip()
                            if not lot or lot == 'None':
                                continue

                            tester     = _cell_str(row, col_tester)
                            flashable  = _cell_str(row, col_flash).lower() == 'да'
                            sw_version = _cell_str(row, col_sw)
                            remote     = _cell_str(row, col_remote)

                            # Характеристики из threaded comments
                            sw_comment = ''
                            try:
                                row_idx = row[0].row if hasattr(row[0], 'row') else None
                                if row_idx:
                                    sw_comment = tc_comments.get((sheet_name, row_idx), '')
                            except Exception:
                                pass

                            remote_id = None
                            if remote:
                                if remote not in remote_cache:
                                    rc = RemoteControl(name=remote)
                                    db.session.add(rc)
                                    db.session.flush()
                                    remote_cache[remote] = rc
                                remote_id = remote_cache[remote].id

                            if (brand.id, model_name, lot) in existing_set:
                                tv_id = existing_map.get((brand.id, model_name, lot))
                                if tv_id:
                                    db.session.query(TVModel).filter_by(id=tv_id).update({
                                        **(({'software_version': sw_version}) if sw_version else {}),
                                        **(({'specifications': sw_comment}) if sw_comment else {}),
                                        **(({'tester_name': tester}) if tester else {}),
                                        **(({'remote_control_id': remote_id}) if remote_id else {}),
                                        'is_flashable': flashable,
                                    })
                                skipped += 1
                                continue

                            tv = TVModel(
                                brand_id=brand.id,
                                launcher_type_id=launcher.id,
                                model=model_name, lot=lot,
                                remote_control_id=remote_id,
                                software_version=sw_version or None,
                                specifications=sw_comment or None,
                                is_flashable=flashable,
                                tester_name=tester or None,
                                tester_id=current_user.id if tester else None,
                                date_added=tc_dates.get((sheet_name, row_idx)) if row_idx else None,
                            )
                            db.session.add(tv)
                            existing_set.add((brand.id, model_name, lot))
                            added += 1

                            if added % BATCH == 0:
                                db.session.commit()

                    wb.close()

                else:
                    import csv
                    content_str = file.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(content_str), delimiter=';')

                    for row in reader:
                        brand_name = row.get('Бренд', '').strip()
                        model_name = row.get('Модель', '').strip()
                        lot        = row.get('Лот', '').strip()
                        if not all([brand_name, model_name, lot]):
                            skipped += 1
                            continue

                        if brand_name not in brand_cache:
                            brand = Brand(name=brand_name)
                            db.session.add(brand)
                            db.session.flush()
                            brand_cache[brand_name] = brand
                        brand = brand_cache[brand_name]

                        if (brand.id, model_name, lot) in existing_set:
                            skipped += 1
                            continue

                        remote_name = row.get('Пульт', '').strip()
                        remote_id = None
                        if remote_name:
                            if remote_name not in remote_cache:
                                rc = RemoteControl(name=remote_name)
                                db.session.add(rc)
                                db.session.flush()
                                remote_cache[remote_name] = rc
                            remote_id = remote_cache[remote_name].id

                        tv = TVModel(
                            brand_id=brand.id, launcher_type_id=launcher.id,
                            model=model_name, lot=lot,
                            remote_control_id=remote_id,
                            software_version=row.get('Версия ПО', '').strip() or None,
                            is_flashable=row.get('Прошивается', '').strip() == 'Да',
                            tester_name=row.get('Тестировщик', '').strip() or current_user.name,
                            tester_id=current_user.id,
                        )
                        db.session.add(tv)
                        existing_set.add((brand.id, model_name, lot))
                        added += 1

                        if added % BATCH == 0:
                            db.session.commit()

                db.session.commit()
                flash(f'Импортировано: {added} моделей, пропущено дублей: {skipped}', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка импорта: {e}', 'error')

            return redirect(url_for('index'))

        return render_template('import_all.html')

    # ── ИНЛАЙН РЕДАКТИРОВАНИЕ ──
    @app.route('/api/inline_edit/<int:id>', methods=['POST'])
    @login_required
    @editor_required
    @csrf.exempt
    def inline_edit(id):
        from flask import jsonify
        model = TVModel.query.get_or_404(id)
        data  = request.get_json()
        field = data.get('field')
        value = data.get('value', '').strip()

        allowed = {'software_version', 'remote_control', 'lot', 'tester_name', 'is_flashable'}
        if field not in allowed:
            return jsonify({'ok': False, 'error': 'Поле недоступно'}), 400

        if field == 'is_flashable':
            setattr(model, field, value in ('true', '1', True))
        elif field == 'remote_control':
            # Ищем пульт по имени, сохраняем FK
            if value:
                remote = RemoteControl.query.filter_by(name=value).first()
                model.remote_control_id = remote.id if remote else None
            else:
                model.remote_control_id = None
        else:
            setattr(model, field, value or None)

        log_action('inline', model, field=field, old_value=None, new_value=value)
        db.session.commit()
        return jsonify({'ok': True, 'value': getattr(model, field)})

    # ── ГЛОБАЛЬНЫЙ ЭКСПОРТ В XLSX (формат Подготовка производства) ──
    @app.route('/export/production')
    @login_required
    def export_production():
        import io
        from flask import Response
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # удаляем пустой лист по умолчанию

        brands = Brand.query.order_by(Brand.name).all()

        # Стили заголовков
        header_font      = Font(name='Calibri', bold=True, size=10)
        header_fill      = PatternFill('solid', fgColor='D9D9D9')
        center_align     = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align       = Alignment(horizontal='left', vertical='center', wrap_text=False)
        thin_border_side = Side(style='thin')
        thin_border      = Border(
            left=thin_border_side, right=thin_border_side,
            top=thin_border_side, bottom=thin_border_side
        )

        for brand in brands:
            models = TVModel.query.filter_by(brand_id=brand.id)                .order_by(TVModel.model, TVModel.lot).all()
            if not models:
                continue

            ws = wb.create_sheet(title=brand.name[:31])

            # Цвет вкладки
            if brand.tab_color:
                ws.sheet_properties.tabColor = brand.tab_color

            # ── Строка 1: заголовок бренда ──
            ws.merge_cells('A1:L1')
            ws['A1'] = f'Подготовка производства {brand.name}'
            ws['A1'].font      = Font(name='Calibri', bold=True, size=12)
            ws['A1'].alignment = center_align

            # ── Строка 2: пустая ──

            # ── Строка 3: заголовки колонок верхний уровень ──
            headers_row3 = [
                'Модель \nтелевизора', 'Лот', 'Слич.вед.', 'SOP',
                'Номер ТП', 'Разработка ТП по сборке', '', 'Разработчик РТП',
                'Программное обеспечение', '', '', ''
            ]
            for col, val in enumerate(headers_row3, 1):
                cell = ws.cell(row=3, column=col, value=val)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border

            # ── Строка 4: подзаголовки ──
            headers_row4 = [
                '', '', '', '', '',
                'Разработчик', 'Статус', '',
                'Шьём', 'Версия ПО', 'STB', 'Macros'
            ]
            for col, val in enumerate(headers_row4, 1):
                cell = ws.cell(row=4, column=col, value=val)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border

            # Объединяем ячейки в шапке (как в оригинале)
            for col in [1, 2, 3, 4, 5, 8]:
                ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
            ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)
            ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=12)

            # ── Данные с 5-й строки ──
            data_font = Font(name='Calibri', size=10, bold=True)
            for tv in models:
                row_data = [
                    tv.model,                                    # A: Модель
                    tv.lot,                                      # B: Лот
                    'Да',                                        # C: Слич.вед.
                    'Да',                                        # D: SOP
                    '',                                          # E: Номер ТП
                    '',                                          # F: Разработчик
                    '',                                          # G: Статус
                    tv.tester_name or '',                        # H: Разработчик РТП
                    'Да' if tv.is_flashable else 'Нет',          # I: Шьём
                    tv.software_version or '',                   # J: Версия ПО
                    tv.remote.name if tv.remote else '',         # K: STB
                    '',                                          # L: Macros
                ]
                ws.append(row_data)
                last_row = ws.max_row
                for col in range(1, 13):
                    cell = ws.cell(row=last_row, column=col)
                    cell.font      = data_font
                    cell.alignment = left_align
                    cell.border    = thin_border

                # Записываем specifications как комментарий в ячейку J (Версия ПО)
                if tv.specifications:
                    from openpyxl.comments import Comment
                    comment = Comment(tv.specifications, tv.tester_name or 'Кабинет технолога')
                    comment.width  = 300
                    comment.height = 120
                    ws.cell(row=last_row, column=10).comment = comment

            # Ширина колонок
            col_widths = [20, 8, 10, 8, 14, 16, 10, 18, 8, 20, 16, 14]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Высота строк шапки
            ws.row_dimensions[3].height = 30
            ws.row_dimensions[4].height = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        from flask import make_response
        resp = make_response(output.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = 'attachment; filename="production_export.xlsx"'
        return resp

    # ── ЭКСПОРТ В EXCEL ──
    @app.route('/export/<int:brand_id>/<int:launcher_id>')
    @login_required
    def export_excel(brand_id, launcher_id):
        import io, csv
        from flask import Response
        brand    = Brand.query.get_or_404(brand_id)
        launcher = LauncherType.query.get_or_404(launcher_id)
        from sqlalchemy.orm import joinedload
        models   = TVModel.query.options(
            joinedload(TVModel.remote),
            joinedload(TVModel.tags),
        ).filter_by(
            brand_id=brand_id, launcher_type_id=launcher_id
        ).order_by(TVModel.date_added.desc()).all()

        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow([
            'Модель', 'Лот', 'Пульт', 'Версия ПО',
            'Прошивается', 'Тестировщик', 'Метки', 'Дата добавления'
        ])
        for m in models:
            writer.writerow([
                m.model,
                m.lot,
                m.remote.name if m.remote else '',
                m.software_version or '',
                'Да' if m.is_flashable else 'Нет',
                m.tester_name or '',
                ', '.join(t.name for t in m.tags),
                m.date_added.strftime('%d.%m.%Y'),
            ])

        output.seek(0)
        filename = f"{brand.name}_{launcher.name}.csv"
        return Response(
            '\ufeff' + output.getvalue(),  # BOM для Excel
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )

    # ── ИМПОРТ ИЗ CSV/EXCEL ──
    @app.route('/import/<int:brand_id>/<int:launcher_id>', methods=['GET', 'POST'])
    @login_required
    def import_models(brand_id, launcher_id):
        brand    = Brand.query.get_or_404(brand_id)
        launcher = LauncherType.query.get_or_404(launcher_id)

        if request.method == 'POST':
            import io, csv
            file = request.files.get('file')
            if not file or not file.filename:
                flash('Выберите файл', 'error')
                return redirect(request.url)

            try:
                content = file.read().decode('utf-8-sig')  # utf-8 с BOM
                reader  = csv.DictReader(io.StringIO(content), delimiter=';')
                added = 0
                skipped = 0
                for row in reader:
                    model_name = row.get('Модель', '').strip()
                    lot        = row.get('Лот', '').strip()
                    if not model_name or not lot:
                        skipped += 1
                        continue
                    existing = TVModel.query.filter_by(
                        brand_id=brand_id, model=model_name, lot=lot
                    ).first()
                    if existing:
                        skipped += 1
                        continue
                    remote_name = row.get('Пульт', '').strip()
                    remote_id = None
                    if remote_name:
                        rc = RemoteControl.query.filter_by(name=remote_name).first()
                        if not rc:
                            rc = RemoteControl(name=remote_name)
                            db.session.add(rc)
                            db.session.flush()
                        remote_id = rc.id
                    tv = TVModel(
                        brand_id=brand_id,
                        launcher_type_id=launcher_id,
                        model=model_name,
                        lot=lot,
                        remote_control_id=remote_id,
                        software_version=row.get('Версия ПО', '').strip() or None,
                        is_flashable=row.get('Прошивается', '').strip() == 'Да',
                        tester_name=row.get('Тестировщик', '').strip() or current_user.name,
                        tester_id=current_user.id,
                    )
                    db.session.add(tv)
                    added += 1

                db.session.commit()
                flash(f'Импортировано: {added} моделей, пропущено: {skipped}', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка импорта: {e}', 'error')

            return redirect(url_for('models_list', brand_id=brand_id, launcher_id=launcher_id))

        return render_template('import.html', brand=brand, launcher=launcher)
